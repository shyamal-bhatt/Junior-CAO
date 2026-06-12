"""
features/normalization/parsers.py
──────────────────────────────────
Blueprint for validating and parsing email/document attachments (e.g., Excel, PDF).
Provides validation architecture mapping:
External Payload -> Attachment Type Validator -> Raw Text Extraction -> Database Insertion
"""

import io
from typing import Dict, Any, List
from app.core.logging import get_logger
from app.features.normalization.chunking import split_text

logger = get_logger(__name__)


class AttachmentParserError(Exception):
    """Base error for attachment parsing."""
    pass



class AttachmentTypeValidator:
    """
    Validates the attachment metadata and routes it based on extension or MIME type.
    """
    
    @staticmethod
    def validate_and_route(filename: str, file_data: bytes) -> str:
        """
        Determines the file type and returns the extracted raw text.
        Part of the flow: External Payload -> Attachment Type Validator
        """
        logger.info(f"[VALIDATION] Validating attachment: {filename} ({len(file_data)} bytes)")
        
        lower_name = filename.lower()
        if lower_name.endswith(".xlsx") or lower_name.endswith(".xls"):
            logger.info(f"[ROUTING] Routing {filename} to Excel parser (openpyxl)")
            return ExcelAttachmentParser.parse(file_data)
        elif lower_name.endswith(".pdf"):
            logger.info(f"[ROUTING] Routing {filename} to PDF parser (PyPDF2)")
            return PDFAttachmentParser.parse(file_data)
        else:
            logger.warning(f"[ROUTING] Unsupported attachment type for {filename}")
            raise AttachmentParserError(f"Unsupported attachment type for file: {filename}")


class ExcelAttachmentParser:
    """
    Parses spreadsheet schedules/tables using openpyxl.
    """
    
    @staticmethod
    def parse(file_data: bytes) -> str:
        """
        Extracts spreadsheet content as structured plaintext.
        Part of the flow: Raw Text Extraction
        """
        try:
            # Dynamically import openpyxl to avoid hard dependency lock at startup
            import openpyxl
            
            logger.info("[PARSING] Executing Excel raw text extraction via openpyxl")
            wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
            text_lines = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text_lines.append(f"--- Sheet: {sheet_name} ---")
                for row in sheet.iter_rows(values_only=True):
                    # Filter out empty rows
                    if any(cell is not None for cell in row):
                        row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
                        text_lines.append(row_str)
            
            raw_text = "\n".join(text_lines)
            logger.info(f"[PARSING] Extraction complete. Extracted {len(raw_text)} chars.")
            return raw_text
        except ImportError:
            logger.warning("openpyxl is not installed. Running mock Excel extraction parser.")
            # Fallback mock parsing
            return f"Mock Excel parsing payload of size {len(file_data)} bytes."
        except Exception as e:
            logger.error(f"Failed to parse Excel file: {e}")
            raise AttachmentParserError(f"Excel parsing error: {e}")


class PDFAttachmentParser:
    """
    Parses PDF text content using PyPDF2.
    """
    
    @staticmethod
    def parse(file_data: bytes) -> str:
        """
        Extracts PDF pages as plaintext.
        Part of the flow: Raw Text Extraction
        """
        try:
            # Dynamically import PyPDF2
            # pyrefly: ignore [missing-import]
            import PyPDF2
            
            logger.info("[PARSING] Executing PDF raw text extraction via PyPDF2")
            reader = PyPDF2.PdfReader(io.BytesIO(file_data))
            text_lines = []
            for idx, page in enumerate(reader.pages):
                text = page.extract_text()
                if text:
                    text_lines.append(f"--- Page {idx + 1} ---")
                    text_lines.append(text)
            
            raw_text = "\n".join(text_lines)
            logger.info(f"[PARSING] Extraction complete. Extracted {len(raw_text)} chars.")
            return raw_text
        except ImportError:
            logger.warning("PyPDF2 is not installed. Running mock PDF extraction parser.")
            return f"Mock PDF parsing payload of size {len(file_data)} bytes."
        except Exception as e:
            logger.error(f"Failed to parse PDF file: {e}")
            raise AttachmentParserError(f"PDF parsing error: {e}")


async def process_attachment_pipeline(
    filename: str,
    file_data: bytes,
    metadata: Dict[str, Any],
    supabase_client: Any,
    embedding_service: Any
) -> Dict[str, Any]:
    """
    Full pipeline execution blueprint:
    External Payload -> Attachment Type Validator -> Raw Text Extraction -> Database Insertion
    """
    logger.info("=========================================")
    logger.info(f"[PIPELINE START] Processing attachment '{filename}'")
    logger.info(f"[TRACE] External Payload -> Attachment Type Validator")
    
    # 1. Validation & Routing & Raw Text Extraction
    try:
        raw_text = AttachmentTypeValidator.validate_and_route(filename, file_data)
    except Exception as e:
        logger.error(f"[PIPELINE FAILED] Validation or extraction failed: {e}")
        raise
        
    logger.info(f"[TRACE] Raw Text Extraction -> Chunking & Vector Generation")
    
    # Chunking
    chunks = split_text(raw_text)
    if not chunks:
        chunks = [f"Attachment: {filename} (empty content)"]
        
    logger.info(f"[TRACE] Generating embeddings for {len(chunks)} chunk(s)")
    embeddings = []
    for chunk in chunks:
        emb = await embedding_service.generate_embedding(chunk)
        embeddings.append(emb)
    
    logger.info(f"[TRACE] Vector Generation -> Database Insertion")
    
    # 3. Database Insertion
    # Insert metadata to raw_documents
    doc_payload = {
        "title": f"Attachment: {filename}",
        "body": raw_text,
        "author": metadata.get("author", "system"),
        "status": metadata.get("status", "processed"),
        "platform": metadata.get("platform", "email"),
        "project_tag": metadata.get("project_tag", "Linkmate"),
        "created_at": metadata.get("created_at")
    }
    
    # Insert row
    logger.info(f"[DATABASE] Writing row to SQL table 'raw_documents'")
    res = supabase_client.table("raw_documents").insert(doc_payload).execute()
    doc_id = res.data[0]["id"]
    
    # Clean old chunks (if any exist) and insert chunks
    logger.info(f"[DATABASE] Inserting {len(chunks)} chunks to table 'document_chunks'.")
    supabase_client.table("document_chunks").delete().eq("document_id", doc_id).execute()
    
    chunk_payloads = [
        {
            "document_id": doc_id,
            "chunk_text": chunk,
            "embedding": emb
        }
        for chunk, emb in zip(chunks, embeddings)
    ]
    supabase_client.table("document_chunks").insert(chunk_payloads).execute()
    
    logger.info(f"[PIPELINE COMPLETE] Successfully processed and stored {filename}")
    logger.info("=========================================")
    
    return {"document_id": doc_id, "text_length": len(raw_text)}
